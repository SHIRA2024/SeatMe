"""
Generate the SeatMe risk register as an Excel workbook (.xlsx).

Deliverable 12. Produces `seatme-risk-register.xlsx` with:
  - "Scales"        : the custom probability and cost scales (NOT 0-1 or dollars)
  - "Risk Register" : >=10 risks with Probability, Cost, Exposure (=P x C), band
  - "Top Risk Mitigation" : mitigation plan for the highest-exposure risk(s)

Usage:
    pip install openpyxl
    python generate_risk_register.py

The same data is also kept as `risks.csv` for reviewers without Excel/openpyxl.
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "seatme-risk-register.xlsx")

# ── Custom scales (deliberately NOT 0-1 and NOT dollars) ────────────────────
PROBABILITY_SCALE = [
    (1, "Rare", "Very unlikely during the project / a typical event"),
    (2, "Unlikely", "Could happen but not expected"),
    (3, "Possible", "Might happen at some point"),
    (4, "Likely", "Expected to happen sometimes"),
    (5, "Almost certain", "Expected to happen most of the time"),
]
COST_SCALE = [
    (1, "Negligible", "Cosmetic / no real effect"),
    (2, "Minor", "Small inconvenience, easy workaround"),
    (3, "Moderate", "Noticeable disruption, some rework"),
    (4, "Major", "Significant outage / data or trust impact"),
    (5, "Severe", "Project- or security-critical failure"),
]

def band(exposure: int) -> str:
    if exposure <= 5:
        return "Low"
    if exposure <= 12:
        return "Medium"
    if exposure <= 19:
        return "High"
    return "Critical"

BAND_FILL = {
    "Low": "C6EFCE",
    "Medium": "FFEB9C",
    "High": "FFC7AE",
    "Critical": "FFC7CE",
}

# ── Risks: (id, category, risk, probability 1-5, cost 1-5, mitigation) ───────
RISKS = [
    ("R01", "Operational", "AWS Academy lab session expires / is reset, wiping all resources", 4, 4,
     "One-command redeploy (redeploy_all.py); keep all infra as code; export DynamoDB data before stopping the lab."),
    ("R02", "Security", "API Gateway has no gateway-level JWT authorizer; auth is enforced in-Lambda instead", 3, 5,
     "Write endpoints and the admin route validate the Cognito token + ownership/group in-Lambda (_common.require_owner); public reads stay open for RSVP. Roadmap: add a Cognito JWT authorizer to the HTTP API."),
    ("R03", "Security", "Weak/known seeded admin password or leaked admin credentials", 3, 5,
     "Force password change on first use; override via SEATME_ADMIN_PASSWORD; keep the admin group small."),
    ("R04", "Delivery", "SNS subscription confirmation friction -> invitations not delivered", 4, 3,
     "Clear UI copy explaining the one-time confirm email; report pending vs sent counts; resend support."),
    ("R05", "Delivery", "SES blocked in lab -> verification emails land in spam / not delivered", 4, 3,
     "Graceful fallback to Cognito built-in mailer; document spam-folder check; resend-code action."),
    ("R06", "Availability", "Single-region (us-east-1) outage takes the whole app down", 2, 4,
     "Accept for MVP; roadmap: multi-region S3 + DynamoDB global tables + Route 53 failover."),
    ("R07", "Data", "Accidental teardown / no backups -> permanent data loss", 3, 4,
     "Confirmation prompts on destructive commands; enable DynamoDB PITR / on-demand backups before teardown."),
    ("R08", "Scalability", "DynamoDB 400 KB item limit exceeded for very large events", 2, 4,
     "Monitor item size; roadmap: split guests into child items keyed by host+guest for big events."),
    ("R09", "Security", "XSS via guest-supplied fields (name, song request)", 2, 4,
     "Frontend escapes all user input with escapeHtml() before DOM insertion; server-side validation."),
    ("R10", "Reliability", "Lambda code regression breaks the API during a redeploy", 3, 3,
     "Idempotent deploys; validate with the OpenAPI spec; roadmap: staged alias + automated smoke tests."),
    ("R11", "Cost", "Cost overrun from traffic spikes or abuse of open endpoints", 2, 3,
     "Pay-per-use scales with usage; roadmap: API throttling/usage plans and AWS Budgets alerts."),
    ("R12", "Privacy", "No-login preview link (?host=) lets anyone with the link view an event's details (read-only)", 3, 3,
     "Editing now requires login (writes authorized in-Lambda), so the link is view-only; treat it as a shareable secret. Roadmap: signed, expiring preview tokens instead of raw email."),
    ("R13", "Config", "Placeholder injection fails -> site can't reach API / Cognito", 2, 4,
     "deploy_frontend.py injects config at upload; 'redeploy to re-inject' documented in troubleshooting."),
]


def style_header(ws, row, ncols):
    fill = PatternFill("solid", fgColor="1F2A44")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def build():
    wb = Workbook()

    # Sheet 1: Scales
    ws = wb.active
    ws.title = "Scales"
    ws["A1"] = "SeatMe — Risk Scales (custom, not 0–1 and not dollars)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A3"] = "Probability scale"
    ws["A3"].font = Font(bold=True)
    ws.append([])  # spacer handled below via explicit rows
    headers = ["Level", "Label", "Meaning"]
    ws.append(headers)
    style_header(ws, ws.max_row, 3)
    for lvl, label, meaning in PROBABILITY_SCALE:
        ws.append([lvl, label, meaning])
    ws.append([])
    ws.append(["Cost / impact scale"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(headers)
    style_header(ws, ws.max_row, 3)
    for lvl, label, meaning in COST_SCALE:
        ws.append([lvl, label, meaning])
    ws.append([])
    ws.append(["Exposure = Probability × Cost  (range 1–25)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Band", "Exposure range"])
    style_header(ws, ws.max_row, 2)
    for b, rng in [("Low", "1–5"), ("Medium", "6–12"), ("High", "13–19"), ("Critical", "20–25")]:
        ws.append([b, rng])
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=BAND_FILL[b])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70

    # Sheet 2: Risk Register
    rr = wb.create_sheet("Risk Register")
    cols = ["ID", "Category", "Risk", "Probability (1-5)", "Cost (1-5)",
            "Exposure (P×C)", "Band", "Mitigation"]
    rr.append(cols)
    style_header(rr, 1, len(cols))
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rows = sorted(RISKS, key=lambda r: r[3] * r[4], reverse=True)
    for rid, cat, risk, p, c, mit in rows:
        exp = p * c
        b = band(exp)
        rr.append([rid, cat, risk, p, c, exp, b, mit])
        r = rr.max_row
        rr.cell(row=r, column=6).fill = PatternFill("solid", fgColor=BAND_FILL[b])
        rr.cell(row=r, column=7).fill = PatternFill("solid", fgColor=BAND_FILL[b])
        for col in range(1, len(cols) + 1):
            cell = rr.cell(row=r, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    widths = [8, 14, 46, 16, 12, 14, 12, 60]
    for i, w in enumerate(widths, start=1):
        rr.column_dimensions[get_column_letter(i)].width = w
    rr.freeze_panes = "A2"

    # Sheet 3: Top Risk Mitigation
    tm = wb.create_sheet("Top Risk Mitigation")
    top = rows[0]
    tid, tcat, trisk, tp, tc, tmit = top
    texp = tp * tc
    tm["A1"] = "Biggest risk — mitigation plan"
    tm["A1"].font = Font(bold=True, size=14)
    lines = [
        ("Risk ID", tid),
        ("Category", tcat),
        ("Risk", trisk),
        ("Probability", f"{tp}/5"),
        ("Cost / impact", f"{tc}/5"),
        ("Exposure", f"{texp} ({band(texp)})"),
    ]
    row = 3
    for k, v in lines:
        tm.cell(row=row, column=1, value=k).font = Font(bold=True)
        tm.cell(row=row, column=2, value=v)
        row += 1
    row += 1
    tm.cell(row=row, column=1, value="Why it is the biggest risk").font = Font(bold=True)
    row += 1
    tm.cell(row=row, column=1, value=(
        "The project runs in an AWS Academy lab whose session/credentials expire and "
        "whose 'Reset' wipes every resource. If that happens during grading the live "
        "link and data disappear, which is both Likely (4) and Major (4)."
    ))
    tm.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 2
    tm.cell(row=row, column=1, value="Mitigation plan").font = Font(bold=True)
    row += 1
    plan = [
        "1. Infrastructure as code: the entire stack redeploys with `python redeploy_all.py` "
        "in minutes, so a wiped environment is fully recoverable.",
        "2. Data export: before stopping the lab, export the DynamoDB `SeatMe` table "
        "(or enable point-in-time recovery / on-demand backup) so event data can be restored.",
        "3. Avoid the lab 'Reset' button; only stop/start the lab to preserve resources.",
        "4. Keep the redeploy idempotent so re-running updates resources in place without "
        "manual cleanup.",
        "5. Document the seeded credentials and demo link so the app can be re-verified "
        "immediately after a redeploy.",
        "6. Roadmap: move to a standard AWS account (no lab expiry) and add DynamoDB "
        "global tables + multi-region S3 for durability.",
    ]
    for p in plan:
        tm.cell(row=row, column=1, value=p).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    tm.cell(row=row, column=1, value="Runner-up risks to watch").font = Font(bold=True)
    row += 1
    for rid, cat, risk, p, c, mit in rows[1:3]:
        tm.cell(row=row, column=1,
                value=f"{rid} ({p*c}, {band(p*c)}): {risk} — {mit}").alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    tm.column_dimensions["A"].width = 110
    tm.column_dimensions["B"].width = 40

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
